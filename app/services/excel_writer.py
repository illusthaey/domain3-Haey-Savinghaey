from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

RAW_COLS = [
    "구분", "성명",
    "국민연금_개인", "국민연금_기관",
    "건강_개인", "건강_기관",
    "요양_개인", "요양_기관",
    "고용_개인실업", "고용_기관실업", "고용_기관고안직능",
    "산재_기관",
    "개인합계", "기관합계", "총계",
]

def _style_ws(ws, df: pd.DataFrame):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(1, j, col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    # body
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(i, j, val)
            c.border = border
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                c.number_format = "#,##0"
                c.alignment = right
            else:
                c.alignment = left if j in (1, 2) else center

    # widths
    for j, col in enumerate(df.columns, start=1):
        sample = [str(col)] + [str(v) for v in df.iloc[:50, j - 1].tolist()]
        w = min(max(10, max(len(s) for s in sample) + 2), 28)
        ws.column_dimensions[get_column_letter(j)].width = w

def write_social_insurance_xlsx(
    raw_df: pd.DataFrame,
    out_path: Path,
    title: str = "사회보험료 통합 산출(업로드 PDF 기반)",
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 요약 먼저
    ws_sum = wb.create_sheet("요약", 0)
    ws_sum["A1"] = title
    ws_sum["A1"].font = Font(bold=True, size=14)

    headers = ["구분", "인원", "개인합계", "기관합계", "총계"]
    for j, h in enumerate(headers, start=1):
        ws_sum.cell(3, j, h).font = Font(bold=True)

    cats = ["근로자4대", "산재만", "방과후(노무제공자)"]
    for idx, cat in enumerate(cats, start=4):
        ws_sum[f"A{idx}"] = cat
        # FILTER는 Excel 365 기준. 구형 엑셀이면 서버에서 시트 자체를 분리생성하도록 바꾸면 됨.
        ws_sum[f"B{idx}"] = f'=IFERROR(ROWS(FILTER(RAW_통합입력!B2:B999, RAW_통합입력!A2:A999="{cat}")),0)'
        ws_sum[f"C{idx}"] = f'=IFERROR(SUM(FILTER(RAW_통합입력!M2:M999, RAW_통합입력!A2:A999="{cat}")),0)'
        ws_sum[f"D{idx}"] = f'=IFERROR(SUM(FILTER(RAW_통합입력!N2:N999, RAW_통합입력!A2:A999="{cat}")),0)'
        ws_sum[f"E{idx}"] = f'=IFERROR(SUM(FILTER(RAW_통합입력!O2:O999, RAW_통합입력!A2:A999="{cat}")),0)'

    gt = 4 + len(cats) + 1
    ws_sum[f"A{gt}"] = "총계"
    ws_sum[f"B{gt}"] = f"=SUM(B4:B{gt-1})"
    ws_sum[f"C{gt}"] = f"=SUM(C4:C{gt-1})"
    ws_sum[f"D{gt}"] = f"=SUM(D4:D{gt-1})"
    ws_sum[f"E{gt}"] = f"=SUM(E4:E{gt-1})"
    for col in "ABCDE":
        ws_sum[f"{col}{gt}"].font = Font(bold=True)

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 16
    ws_sum.column_dimensions["E"].width = 16

    # RAW 시트
    ws_raw = wb.create_sheet("RAW_통합입력")
    _style_ws(ws_raw, raw_df)
    tab = Table(displayName="T_RAW", ref=f"A1:{get_column_letter(raw_df.shape[1])}{raw_df.shape[0]+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_raw.add_table(tab)

    # 분류 시트들(필터 기반)
    def add_filter_sheet(sheet_name: str, category: str):
        ws = wb.create_sheet(sheet_name)
        # header만 세팅
        _style_ws(ws, raw_df.iloc[0:0].copy())  # 빈 df로 헤더/폭
        last_row = raw_df.shape[0] + 1
        ws["A2"] = f'=IFERROR(FILTER(RAW_통합입력!A2:O{last_row}, RAW_통합입력!A2:A{last_row}="{category}"), "")'

    add_filter_sheet("근로자4대", "근로자4대")
    add_filter_sheet("산재만(교육공무직)", "산재만")
    add_filter_sheet("방과후(노무제공자)", "방과후(노무제공자)")

    wb.save(str(out_path))
